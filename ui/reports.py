import streamlit as st
import pandas as pd
import io

# LIBRERÍAS VISUALES PESADAS
from streamlit_lottie import st_lottie
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from st_aggrid import AgGrid, GridOptionsBuilder, ColumnsAutoSizeMode
from streamlit_echarts import st_echarts
from streamlit_elements import elements, mui

from ui.design import load_lottieurl

def render_dashboard_bi(historial):
    st.markdown("### 📊 Business Intelligence & Exportación")
    
    if not historial:
        st.info("ℹ️ Realiza al menos una simulación para activar el dashboard de análisis.")
        lottie_logistics = load_lottieurl("https://assets9.lottiefiles.com/packages/lf20_s2l79gze.json")
        if lottie_logistics: 
            st_lottie(lottie_logistics, height=300, key="empty_state")
        return

    df_h = pd.DataFrame(historial)
    avg_viab = df_h['Viabilidad (Res)'].mean()

    # ---------------------------------------------------------
    # TARJETAS DE MÉTRICAS (MATERIAL UI)
    # ---------------------------------------------------------
    with elements("dashboard_metrics"):
        with mui.Stack(direction="row", spacing=3, sx={"mb": 4, "mt": 2}):
            with mui.Card(elevation=0, sx={"flex": 1, "p": 3, "borderRadius": 4, "border": "1px solid #E1E5F2", "transition": "0.3s", "&:hover": {"boxShadow": "0 8px 24px rgba(46,91,255,0.1)", "transform": "translateY(-2px)"}}):
                mui.Typography("Total SKU Analizados", variant="subtitle2", sx={"color": "#6B778C", "fontWeight": 600, "mb": 1})
                mui.Typography(f"{len(df_h)}", variant="h3", sx={"color": "#091E42", "fontWeight": 800})
            
            color_viab = "#00E676" if avg_viab >= 1.5 else ("#FFD166" if avg_viab >= 1.2 else "#FF4B4B")
            with mui.Card(elevation=0, sx={"flex": 1, "p": 3, "borderRadius": 4, "border": "1px solid #E1E5F2", "transition": "0.3s", "&:hover": {"boxShadow": "0 8px 24px rgba(46,91,255,0.1)", "transform": "translateY(-2px)"}}):
                mui.Typography("Viabilidad Promedio", variant="subtitle2", sx={"color": "#6B778C", "fontWeight": 600, "mb": 1})
                with mui.Stack(direction="row", alignItems="baseline", spacing=1):
                    mui.Typography(f"{avg_viab:.2f}x", variant="h3", sx={"color": color_viab, "fontWeight": 800})
                    mui.Typography("Meta: 1.5x", variant="caption", sx={"color": "#A5ADBA", "fontWeight": 500})
                    
            with mui.Card(elevation=0, sx={"flex": 1, "p": 3, "borderRadius": 4, "border": "1px solid #E1E5F2", "transition": "0.3s", "&:hover": {"boxShadow": "0 8px 24px rgba(46,91,255,0.1)", "transform": "translateY(-2px)"}}):
                mui.Typography("Costo Unitario Promedio", variant="subtitle2", sx={"color": "#6B778C", "fontWeight": 600, "mb": 1})
                mui.Typography(f"${df_h['Costo Unitario (Res)'].mean():,.0f}", variant="h3", sx={"color": "#2E5BFF", "fontWeight": 800})

    # ---------------------------------------------------------
    # SELECTOR DE VISTAS (Solución al Bug de ECharts en Tabs)
    # ---------------------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    
    vista_actual = st.radio(
        "Navegación del Reporte:",
        ["📊 Visión General", "🫧 Riesgo e Inversión (Avanzado)"],
        horizontal=True,
        label_visibility="collapsed"
    )

    # --- VISTA 1: VISIÓN GENERAL ---
    if vista_actual == "📊 Visión General":
        st.markdown("<br>", unsafe_allow_html=True)
        g1, g2 = st.columns(2)
        
        with g1:
            st.markdown("<h5 style='text-align:center; color:#091E42;'>💰 Balance Financiero: Costo vs Ingreso</h5>", unsafe_allow_html=True)
            productos = df_h['Producto'].tolist()
            costos = df_h['Costo Unitario (Res)'].round(0).tolist()
            ingresos = df_h['Ingreso ML (Res)'].round(0).tolist()
            
            option_bar = {
                "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
                "legend": {"data": ["Costo Unitario", "Ingreso Neto"], "bottom": 0},
                "grid": {"left": "3%", "right": "4%", "bottom": "15%", "containLabel": True},
                "xAxis": {"type": "category", "data": productos, "axisLine": {"show": False}},
                "yAxis": {"type": "value", "splitLine": {"lineStyle": {"type": "dashed", "color": "#E1E5F2"}}},
                "color": ["#FF4B4B", "#00E676"],
                "series": [
                    {"name": "Costo Unitario", "type": "bar", "data": costos, "itemStyle": {"borderRadius": [6, 6, 0, 0]}, "barGap": "15%"},
                    {"name": "Ingreso Neto", "type": "bar", "data": ingresos, "itemStyle": {"borderRadius": [6, 6, 0, 0]}}
                ]
            }
            st_echarts(options=option_bar, height="350px", key="bar_chart")
            
        with g2:
            st.markdown("<h5 style='text-align:center; color:#091E42;'>⚖️ Tacómetro de Rentabilidad Promedio</h5>", unsafe_allow_html=True)
            option_gauge = {
                "tooltip": {"formatter": "{a} <br/>{b} : {c}x"},
                "series": [
                    {
                        "name": "Rentabilidad",
                        "type": "gauge",
                        "min": 0, "max": 3, "splitNumber": 3,
                        "axisLine": {"lineStyle": {"width": 18, "color": [[0.4, "#FF4B4B"], [0.5, "#FFD166"], [1, "#00E676"]]}},
                        "pointer": {"itemStyle": {"color": "auto"}},
                        "axisTick": {"distance": -20, "length": 8, "lineStyle": {"color": "#fff", "width": 2}},
                        "splitLine": {"distance": -20, "length": 20, "lineStyle": {"color": "#fff", "width": 3}},
                        "axisLabel": {"color": "inherit", "distance": 30, "fontSize": 12},
                        "detail": {"valueAnimation": True, "formatter": "{value}x", "color": "inherit", "fontSize": 28, "fontWeight": "bold", "padding": [40, 0, 0, 0]},
                        "data": [{"value": round(avg_viab, 2), "name": "Ratio"}]
                    }
                ]
            }
            st_echarts(options=option_gauge, height="350px", key="gauge_chart")

    # --- VISTA 2: GRÁFICAS AVANZADAS ---
    elif vista_actual == "🫧 Riesgo e Inversión (Avanzado)":
        st.markdown("<br>", unsafe_allow_html=True)
        c_adv1, c_adv2 = st.columns(2)
        
        with c_adv1:
            st.markdown("<h5 style='text-align:center; color:#091E42;'>🫧 Cuadrante Mágico (Viabilidad vs Costo)</h5>", unsafe_allow_html=True)
            max_ingreso = float(df_h['Ingreso ML (Res)'].max()) if not df_h.empty else 1.0
            scatter_data = []
            for _, row in df_h.iterrows():
                costo_nativo = float(row['Costo Unitario (Res)'])
                viab_nativa = float(row['Viabilidad (Res)'])
                ingreso_nativo = float(row['Ingreso ML (Res)'])
                b_size = float((ingreso_nativo / max_ingreso) * 40 + 15) if max_ingreso > 0 else 20.0
                
                # LA SOLUCIÓN: Texto limpio, elegante y sin HTML para evitar el bloqueo del escudo de ECharts
                texto_limpio = f"{str(row['Producto'])}  •  💰 ${costo_nativo:,.0f} COP  •  ⚖️ {viab_nativa:.2f}x"
                
                scatter_data.append({
                    "name": texto_limpio,
                    "value": [round(costo_nativo, 0), round(viab_nativa, 2)],
                    "symbolSize": b_size,
                    "itemStyle": {"opacity": 0.8}
                })
            
            option_scatter = {
                "tooltip": {"trigger": "item", "formatter": "{b}"},
                "xAxis": {"name": "Costo Unitario (COP)", "type": "value", "splitLine": {"lineStyle": {"type": "dashed", "color": "#E1E5F2"}}},
                "yAxis": {"name": "Viabilidad (x)", "type": "value", "splitLine": {"lineStyle": {"type": "dashed", "color": "#E1E5F2"}}},
                "series": [{"type": "scatter", "data": scatter_data, "itemStyle": {"color": "#2E5BFF"}}]
            }
            st_echarts(options=option_scatter, height="350px", key="scatter_chart_fix_final")
            
        with c_adv2:
            st.markdown("<h5 style='text-align:center; color:#091E42;'>🍩 Distribución de Inversión por Método</h5>", unsafe_allow_html=True)
            metodos_data = df_h.groupby('Método')['Costo Unitario (Res)'].sum().reset_index()
            pie_data = [{"value": float(row['Costo Unitario (Res)']), "name": str(row['Método'])} for _, row in metodos_data.iterrows()]
            
            option_donut = {
                "tooltip": {"trigger": "item", "formatter": "{b}: ${c} ({d}%)"},
                "legend": {"bottom": "0%", "left": "center"},
                "series": [
                    {
                        "name": "Inversión por Método", "type": "pie", "radius": ["40%", "70%"],
                        "avoidLabelOverlap": False,
                        "itemStyle": {"borderRadius": 10, "borderColor": "#ffffff", "borderWidth": 2},
                        "label": {"show": False, "position": "center"},
                        "emphasis": {"label": {"show": True, "fontSize": "20", "fontWeight": "bold"}},
                        "labelLine": {"show": False},
                        "data": pie_data,
                        "color": ["#2E5BFF", "#00C6FF", "#FFD166", "#FF4B4B"]
                    }
                ]
            }
            st_echarts(options=option_donut, height="350px", key="donut_chart_fix")

    # ---------------------------------------------------------
    # TABLA DE DATOS CON AgGrid
    # ---------------------------------------------------------
    st.markdown("<br><h4 style='color: #091E42; font-weight: 600;'>📑 Registro Detallado de Simulaciones</h4>", unsafe_allow_html=True)
    
    gb = GridOptionsBuilder.from_dataframe(df_h)
    gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=10)
    gb.configure_side_bar() 
    gb.configure_column("Costo Unitario (Res)", type=["numericColumn"], valueFormatter="value != undefined ? '$' + value.toLocaleString('es-CO', {maximumFractionDigits: 0}) : ''")
    gb.configure_column("Ingreso ML (Res)", type=["numericColumn"], valueFormatter="value != undefined ? '$' + value.toLocaleString('es-CO', {maximumFractionDigits: 0}) : ''")
    gb.configure_column("Viabilidad (Res)", type=["numericColumn"], valueFormatter="value != undefined ? value.toFixed(2) + 'x' : ''")
    gb.configure_default_column(editable=False, groupable=True)
    
    AgGrid(df_h, gridOptions=gb.build(), enable_enterprise_modules=False, allow_unsafe_jscode=True, theme='alpine', columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS, height=300)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # ---------------------------------------------------------
    # EXPORTACIÓN EXCEL Y PURGA
    # ---------------------------------------------------------
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as wr:
        df_h.to_excel(wr, index=False, sheet_name='Reporte Gerencial')
        ws = wr.sheets['Reporte Gerencial']
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="2E5BFF", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
             ws.column_dimensions[col[0].column_letter].width = 20
        ws.conditional_formatting.add(f"E2:E{len(df_h)+1}", CellIsRule(operator='greaterThan', formula=['1.5'], fill=PatternFill(start_color="C6EFCE", fill_type="solid")))
        ws.conditional_formatting.add(f"E2:E{len(df_h)+1}", CellIsRule(operator='lessThan', formula=['1.2'], fill=PatternFill(start_color="FFC7CE", fill_type="solid")))
        
    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        st.download_button("📥 Descargar Reporte Excel", buffer.getvalue(), "Reporte_ImportPro.xlsx", type="primary")
    with col_btn2:
        if st.button("🗑️ Purgar Datos del Historial"):
            st.session_state['historial'] = []
            st.rerun()
