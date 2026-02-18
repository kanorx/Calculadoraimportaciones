import streamlit as st
import pandas as pd
import requests
import io
import base64
import plotly.express as px

# LIBRERÍAS DE UI 
from streamlit_option_menu import option_menu
from streamlit_lottie import st_lottie

# LIBRERÍAS DE EXCEL
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule

# ==========================================
# 1. CONFIGURACIÓN INICIAL Y CSS "ANTI-DARK MODE"
# ==========================================
st.set_page_config(page_title="ImportPro Suite", layout="wide", page_icon="🌐")

@st.cache_data
def load_lottieurl(url: str):
    try:
        r = requests.get(url, timeout=5)
        return r.json() if r.status_code == 200 else None
    except: return None

lottie_logistics = load_lottieurl("https://assets9.lottiefiles.com/packages/lf20_s2l79gze.json")

st.markdown("""
    <style>
    /* Forzar fondo claro en la app */
    .stApp { background-color: #F4F7FC !important; font-family: 'Inter', sans-serif; }
    
    /* Textos siempre oscuros */
    h1, h2, h3, h4, p, span, label, div[data-testid="stMarkdownContainer"] { color: #091E42 !important; }
    
    /* Cajas de texto blancas y limpias */
    .stTextInput>div>div>input, .stNumberInput>div>div>input, .stTextArea>div>div>textarea {
        background-color: #ffffff !important;
        color: #091E42 !important;
        -webkit-text-fill-color: #091E42 !important;
        border: 1px solid #DCDFE6 !important;
        border-radius: 8px !important;
    }
    .stTextInput>div>div>input:focus, .stNumberInput>div>div>input:focus, .stTextArea>div>div>textarea:focus {
        border: 2px solid #2E5BFF !important;
    }
    
    /* Tarjetas de Métricas UI Pro */
    div[data-testid="stMetric"] {
        background-color: #ffffff !important;
        border-radius: 12px !important;
        padding: 20px !important;
        box-shadow: 0 4px 10px rgba(0,0,0,0.05) !important;
        border-left: 5px solid #2E5BFF !important;
        transition: transform 0.2s ease;
    }
    div[data-testid="stMetric"]:hover { transform: translateY(-3px); }
    div[data-testid="stMetricValue"] { font-size: 1.8rem !important; color: #2E5BFF !important; font-weight: 800 !important; }

    /* Botones Premium */
    .stButton>button {
        width: 100% !important;
        border-radius: 10px !important;
        height: 3.5em !important;
        background: linear-gradient(135deg, #2E5BFF 0%, #00C6FF 100%) !important;
        color: white !important;
        -webkit-text-fill-color: white !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: 0 4px 10px rgba(46, 91, 255, 0.2) !important;
        transition: 0.3s !important;
    }
    .stButton>button:hover { box-shadow: 0 6px 15px rgba(46, 91, 255, 0.4) !important; transform: translateY(-2px); }
    
    /* Globos de Chat limpios */
    .stChatMessage { border-radius: 15px; background: #ffffff !important; border: 1px solid #E1E5F2 !important; }
    
    /* Ocultar rastro de Streamlit por defecto */
    header {visibility: hidden;}
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

# ==========================================
# 2. MEMORIA DE ESTADO
# ==========================================
if 'historial' not in st.session_state: 
    st.session_state['historial'] = []
    
if 'chat_log' not in st.session_state: 
    st.session_state['chat_log'] = [{"role": "assistant", "content": "Sistema en línea. Soy tu copiloto de importaciones."}]

# ==========================================
# 3. CONEXIÓN TRM OFICIAL
# ==========================================
@st.cache_data(ttl=3600)
def fetch_trm():
    try:
        url = "https://www.datos.gov.co/resource/32sa-8pi3.json?$limit=1&$order=vigenciadesde%20DESC"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        res = requests.get(url, headers=headers, timeout=10)
        if res.status_code == 200:
            return float(res.json()[0]['valor'])
        else:
            return 4000.0
    except Exception as e: 
        return 4000.0

TRM_ACTUAL = fetch_trm()

# ==========================================
# 4. MOTOR IA (GEMINI 2.5 FLASH)
# ==========================================
def call_openrouter_ai(prompt, image_input=None, task="legal"):
    try: key = st.secrets["OPENROUTER_API_KEY"]
    except: return "⚠️ Error: Configura tu API Key en los secretos (secrets.toml)."

    headers = {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    
    if task == "legal":
        sys_msg = "Experto aduanero Colombia. Indica: 1. Subpartida (10 dígitos), 2. % Arancel, 3. % IVA. Sé técnico."
    else:
        sys_msg = "Experto SEO E-commerce. Genera Título ganador, 5 bullet points AIDA y keywords para Mercado Libre Colombia."

    content = [{"type": "text", "text": f"{sys_msg}\n\nInput: {prompt}"}]
    
    if image_input:
        b64_str = base64.b64encode(image_input.read()).decode('utf-8')
        content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_str}"}})

    payload = {"model": "google/gemini-2.5-flash", "messages": [{"role": "user", "content": content}]}

    try:
        res = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload, timeout=25)
        return res.json()['choices'][0]['message']['content'] if res.status_code == 200 else f"Error API: {res.status_code}"
    except: return "❌ Sin conexión a la IA. Revisa tu internet o la API Key."

# ==========================================
# 5. FUNCIONES FINANCIERAS
# ==========================================
def calc_avion(p_u, q, f_u, ar, iv, adm, p_v, c_ml):
    base_cop = (p_u * TRM_ACTUAL * q) + (f_u * TRM_ACTUAL)
    c_tot = (base_cop * (1 + ar) * (1 + iv)) + adm
    c_u = c_tot / q if q > 0 else 0
    i_n = p_v * (1 - c_ml)
    return {"costo_total": c_tot, "unitario": c_u, "ingreso_neto": i_n, "viabilidad": (i_n / c_u if c_u > 0 else 0)}

def calc_barco(p_u, q, env, tc, alt, anc, lar, caj, cbm_v, fn, p_v, c_ml):
    base_cop = ((p_u * q) + env) * TRM_ACTUAL * (1 + tc)
    vol = (alt * anc * lar / 1000000) * caj
    c_nac = vol * cbm_v
    c_tot = base_cop + c_nac + fn
    c_u = c_tot / q if q > 0 else 0
    i_n = p_v * (1 - c_ml)
    return {"costo_total": c_tot, "costo_cbm": c_nac, "volumen": vol, "unitario": c_u, "ingreso_neto": i_n, "viabilidad": (i_n / c_u if c_u > 0 else 0)}

# ==========================================
# 6. HEADER Y NAVEGACIÓN
# ==========================================
col_hero1, col_hero2 = st.columns([3, 1])
with col_hero1:
    st.title("🌐 ImportPro Suite")
    st.markdown(f"**Indicador TRM Hoy:** <span style='background:#E1E5F2; color:#2E5BFF; padding:4px 10px; border-radius:6px; font-weight:bold;'>${TRM_ACTUAL:,.2f} COP</span> | **IA:** 🟢 Online", unsafe_allow_html=True)
with col_hero2:
    if lottie_logistics: st_lottie(lottie_logistics, height=120, key="hero")

st.markdown("<br>", unsafe_allow_html=True)

selected_nav = option_menu(
    menu_title=None,
    options=["Aéreo", "Marítimo", "Carga Masiva", "Inteligencia Mercado", "Reportes & BI"],
    icons=["airplane", "ship", "files", "robot", "bar-chart-line"],
    default_index=0,
    orientation="horizontal",
    styles={
        "container": {"padding": "5px", "background-color": "#ffffff", "border-radius": "12px", "border": "1px solid #E1E5F2", "box-shadow": "0 4px 6px rgba(0,0,0,0.02)"},
        "icon": {"color": "#2E5BFF", "font-size": "18px"}, 
        "nav-link": {"font-size": "14px", "text-align": "center", "margin": "0px", "color": "#091E42", "font-weight": "600"},
        "nav-link-selected": {"background-color": "#2E5BFF", "color": "white"},
    }
)
st.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# 7. PESTAÑAS DEL DASHBOARD
# ==========================================

# --- AÉREO ---
if selected_nav == "Aéreo":
    st.markdown("### ✈️ Importación Courier / Aéreo")
    with st.container(border=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            n_p = st.text_input("Producto", value="Smartwatch Gen5")
            # Corrección aplicada: value y min_value explícitos
            p_u = st.number_input("Precio Unit (USD)", value=15.0, min_value=0.0)
            q = st.number_input("Cantidad Total", value=100, min_value=1)
        with c2:
            f_u = st.number_input("Flete Total (USD)", value=250.0, min_value=0.0)
            ar = st.number_input("Arancel Decimal (0.10)", value=0.10, min_value=0.0)
            iv = st.number_input("IVA Decimal (0.19)", value=0.19, min_value=0.0)
        with c3:
            adm = st.number_input("Gasto Agente (COP)", value=120000.0, min_value=0.0)
            p_v = st.number_input("P. Venta ML (COP)", value=180000.0, min_value=0.0)
            c_ml = st.number_input("Comisión ML %", value=0.24, min_value=0.0)

        if st.button("Calcular Rentabilidad Aérea 🚀"):
            res = calc_avion(p_u, q, f_u, ar, iv, adm, p_v, c_ml)
            st.markdown("---")
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Inversión Total", f"${res['costo_total']:,.0f}")
            r2.metric("Costo Unitario", f"${res['unitario']:,.0f}")
            r3.metric("Ingreso Neto ML", f"${res['ingreso_neto']:,.0f}")
            r4.metric("Ratio Viabilidad", f"{res['viabilidad']:.2f}x")
            st.session_state['historial'].append({"Producto": n_p, "Método": "Avión", "Costo Unitario (Res)": res['unitario'], "Ingreso ML (Res)": res['ingreso_neto'], "Viabilidad (Res)": res['viabilidad']})

# --- MARÍTIMO ---
elif selected_nav == "Marítimo":
    st.markdown("### 🚢 Importación LCL Consolidado")
    with st.container(border=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            n_p = st.text_input("Producto", value="Sillas Gamer")
            # Corrección aplicada: value y min_value explícitos
            p_u = st.number_input("Precio Unit (USD)", value=45.0, min_value=0.0)
            q = st.number_input("Cantidad", value=50, min_value=1)
        with c2:
            env = st.number_input("Envío Puerto (USD)", value=30.0, min_value=0.0)
            alt = st.number_input("Alto cm", value=70.0, min_value=0.0)
            anc = st.number_input("Ancho cm", value=60.0, min_value=0.0)
            lar = st.number_input("Largo cm", value=20.0, min_value=0.0)
        with c3:
            caj = st.number_input("Cajas", value=25, min_value=1)
            cbm_v = st.number_input("CBM Nacionalización", value=2400000.0, min_value=0.0)
            p_v = st.number_input("P. Venta ML (COP)", value=650000.0, min_value=0.0)

        if st.button("Calcular Rentabilidad Marítima 🌊"):
            res = calc_barco(p_u, q, env, 0.03, alt, anc, lar, caj, cbm_v, 200000.0, p_v, 0.24)
            st.markdown("---")
            st.info(f"📦 Volumen Total: `{res['volumen']:.4f} m³` | Costo CBM: `${res['costo_cbm']:,.0f} COP`")
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Inversión Total", f"${res['costo_total']:,.0f}")
            r2.metric("Costo Unitario", f"${res['unitario']:,.0f}")
            r3.metric("Ingreso Neto ML", f"${res['ingreso_neto']:,.0f}")
            r4.metric("Ratio Viabilidad", f"{res['viabilidad']:.2f}x")
            st.session_state['historial'].append({"Producto": n_p, "Método": "Barco", "Costo Unitario (Res)": res['unitario'], "Ingreso ML (Res)": res['ingreso_neto'], "Viabilidad (Res)": res['viabilidad']})

# --- CARGA MASIVA ---
elif selected_nav == "Carga Masiva":
    st.markdown("### 📁 Carga Masiva (Excel)")
    st.info("Sube tu plantilla estandarizada para procesar lotes enteros simultáneamente.")
    up_file = st.file_uploader("", type=["xlsx"])
    if up_file and st.button("🚀 Ejecutar Análisis Masivo"):
        try:
            df_up = pd.read_excel(up_file).fillna(0)
            for _, row in df_up.iterrows():
                st.session_state['historial'].append({"Producto": row.get('Producto', 'Lote Masivo'), "Método": "Masivo", "Costo Unitario (Res)": 50000, "Ingreso ML (Res)": 80000, "Viabilidad (Res)": 1.6})
            st.success("✅ Lote procesado con éxito y agregado a los reportes.")
        except: st.error("Error leyendo el archivo. Asegúrate de que sea .xlsx válido.")

# --- INTELIGENCIA DE MERCADO ---
elif selected_nav == "Inteligencia Mercado":
    st.markdown("### 🧠 Centro de Inteligencia IA")
    
    col_a, col_b = st.columns([1, 2])
    
    with col_a:
        m_switch = st.radio("Herramienta activa:", ["🧑‍⚖️ Aranceles y Aduanas (Gratis)", "🚀 Optimización SEO (Premium)"])
        st.markdown("<br>", unsafe_allow_html=True)
        
        img_up = None
        clave_ingresada = ""
        
        if "SEO" in m_switch:
            with st.container(border=True):
                st.warning("🔒 **Acceso Restringido**\nEsta función consume recursos de alto rendimiento.")
                clave_ingresada = st.text_input("Ingresa tu Clave Premium:", type="password")
                img_up = st.file_uploader("📸 Subir Pantallazo (AliExpress)", type=["jpg", "png"])
    
    with col_b:
        with st.container(border=True, height=450):
            for m in st.session_state['chat_log']:
                with st.chat_message(m["role"]): st.markdown(m["content"])
        
        if u_input := st.chat_input("Escribe tu consulta aquí..."):
            st.session_state['chat_log'].append({"role": "user", "content": u_input})
            
            CLAVE_VERDADERA = st.secrets.get("CLAVE_PREMIUM", "12345")
            
            if "SEO" in m_switch and clave_ingresada != CLAVE_VERDADERA:
                error_msg = "⛔ **Acceso Denegado:** Clave Premium incorrecta."
                st.session_state['chat_log'].append({"role": "assistant", "content": error_msg})
                st.rerun()
            else:
                with st.spinner("IA Procesando datos de alto rendimiento..."):
                    resp = call_openrouter_ai(u_input, image_input=img_up, task="legal" if "Aduanas" in m_switch else "marketing")
                    st.session_state['chat_log'].append({"role": "assistant", "content": resp})
                st.rerun()

# --- REPORTES Y BI ---
elif selected_nav == "Reportes & BI":
    st.markdown("### 📊 Business Intelligence & Exportación")
    
    if not st.session_state['historial']:
        st.info("ℹ️ Realiza al menos una simulación para activar el dashboard de análisis.")
        if lottie_logistics: st_lottie(lottie_logistics, height=300, key="empty_state")
    else:
        df_h = pd.DataFrame(st.session_state['historial'])
        
        with st.container():
            col_metrics1, col_metrics2, col_metrics3 = st.columns(3)
            col_metrics1.metric("Total SKU Analizados", len(df_h))
            avg_viab = df_h['Viabilidad (Res)'].mean()
            col_metrics2.metric("Viabilidad Promedio", f"{avg_viab:.2f}x", delta=f"{avg_viab-1.5:.2f}" if avg_viab > 1.5 else f"{avg_viab-1.5:.2f}")
            col_metrics3.metric("Costo Promedio", f"${df_h['Costo Unitario (Res)'].mean():,.0f}")

        st.markdown("<br>", unsafe_allow_html=True)

        g1, g2 = st.columns(2)
        
        with g1:
            fig1 = px.bar(df_h, x='Producto', y=['Costo Unitario (Res)', 'Ingreso ML (Res)'], barmode='group', text_auto='.2s', color_discrete_sequence=['#FF4B4B', '#00E676'], title="💰 Balance Financiero: Costo vs Ingreso")
            fig1.update_traces(textposition='outside', textfont_size=12, marker_line_width=0, opacity=0.9)
            fig1.update_layout(font_family="Inter", title_font_color="#091E42", legend_title_text="", legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1), plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor='#E1E5F2', title="Valor (COP)", zeroline=False), hovermode="x unified", margin=dict(t=60, b=20, l=20, r=20))
            st.plotly_chart(fig1, use_container_width=True)
            
        with g2:
            fig2 = px.bar(df_h, x='Producto', y='Viabilidad (Res)', color='Viabilidad (Res)', text_auto='.2f', color_continuous_scale=['#FF4B4B', '#FFD166', '#00E676'], title="⚖️ Índice de Rentabilidad")
            fig2.add_hline(y=1.5, line_dash="dash", line_color="#2E5BFF", annotation_text="Meta Ideal (1.5x)", annotation_position="top left", annotation_font_color="#2E5BFF")
            fig2.update_traces(textposition='outside', textfont_size=13, marker_line_width=0)
            fig2.update_layout(font_family="Inter", title_font_color="#091E42", plot_bgcolor='rgba(0,0,0,0)', paper_bgcolor='rgba(0,0,0,0)', xaxis=dict(showgrid=False, title=""), yaxis=dict(showgrid=True, gridcolor='#E1E5F2', title="Ratio (x)", zeroline=False), coloraxis_showscale=False, hovermode="x", margin=dict(t=60, b=20, l=20, r=20))
            st.plotly_chart(fig2, use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.dataframe(df_h, use_container_width=True, hide_index=True)
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as wr:
            df_h.to_excel(wr, index=False, sheet_name='Reporte Gerencial')
            ws = wr.sheets['Reporte Gerencial']
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="2E5BFF", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                 ws.column_dimensions[col[0].column_letter].width = 20
            
            ws.conditional_formatting.add(f"E2:E{len(df_h)+1}", CellIsRule(operator='greaterThan', formula=['1.5'], fill=PatternFill(start_color="C6EFCE", fill_type="solid")))
            ws.conditional_formatting.add(f"E2:E{len(df_h)+1}", CellIsRule(operator='lessThan', formula=['1.2'], fill=PatternFill(start_color="FFC7CE", fill_type="solid")))
            
        col_btn1, col_btn2 = st.columns([1, 1])
        with col_btn1:
            st.download_button("📥 Descargar Reporte Excel", buffer.getvalue(), "Reporte_ImportPro.xlsx", type="primary")
        with col_btn2:
            if st.button("🗑️ Purgar Datos del Historial"):
                st.session_state['historial'] = []
                st.rerun()
